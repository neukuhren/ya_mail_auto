"""Модуль для работы с excel файлами"""
from xlsxwriter import Workbook
from openpyxl import load_workbook

import logging # Импортируем библиотеку для безопасного хранения логов
logger = logging.getLogger(__name__)

from config import LIST_WITH_COLUMNS_COMBINED





def write_to_excel(list_with_data, keyword='combined'):
    """Производит запись в xlsx файл"""
    wb = Workbook(f'{keyword}.xlsx')
    ws = wb.add_worksheet(f'{keyword}') # По дефоолту 'sheet 1'
    ordered_list = LIST_WITH_COLUMNS_COMBINED
    first_row = 0
    for header in ordered_list:
        col = ordered_list.index(header) # Порядок столбцов
        ws.write(first_row, col, header) # Первая строка - заголовок

    row = 1
    for dict_ in list_with_data:
        for _key,_value in dict_.items():
            col=ordered_list.index(_key)
            ws.write(row,col,_value)
        row += 1 # Перевод на следующую строку
    wb.close()


def open_xlsx_file_and_return_active_sheet(file_name='combined.xlsx'):
    """Открывает файл xlsx и возвращает книгу и текущий лист.
    """
    # Найти целевой файл и рабочий лист
    try:
        workbook = load_workbook(file_name) 
        worksheet = workbook.active  # делаем единственный лист активным 
        if worksheet:
            logger.debug(f'Рабочий лист {worksheet} найден')
        else:
            logger.debug(f'Рабочий лист {worksheet} не найден')
    except Exception as exc:
        logger.critical(f'КРИТИЧЕСКАЯ ОШИБКА: xlsx файл или лист. Получение данных невозможно. {exc}')
    return workbook, worksheet



def read_xslx_with_filter_col1_col2(worksheet, num_col_filter, filter_value, num_col_target) -> list:
    """Возвращает список значений столбца, только тех строк, которые удовлетворяют условию.

    Args:
        worksheet - активный рабочий лист xslx, 
        num_col_filter - номер столбца, по которому фильтровать,
        filter_value - значение, по которому фильтровать (условие),
        num_col_target - номер целевого столбца.
    Return:
        list_with_filtered_values - список с отфильтрованными значениями.
    """
    list_with_filtered_values = []  # список с отфильтрованными значениям (куки)
    if worksheet:  # Если активный лист получены
        current_row = 2  # С какой строки читать xlsx файл
        list_with_filtered_values.clear()

        # Пройтись со второй строки до первой пустой
        try: 
            # Находим ячейку в солбце 'Status'
            filter_cell = worksheet.cell(row=current_row, column=num_col_filter)
            while filter_cell.value is not None:
                try:
                    if filter_cell.value == filter_value:  # Если статус ок
                        target_cell = worksheet.cell(row=current_row, column=num_col_target)
                        list_with_filtered_values.append(target_cell.value)
                except:
                    logger.error(f'При проверке значений в {current_row} возникла ошибка')
                finally:
                    current_row += 1
                    filter_cell = worksheet.cell(row=current_row, column=num_col_filter)  # Переход на следующую строку
                    # pause = input('Нажмите любую клавишу для продолжения работы: ')
            else:
                logger.debug(f'Обнаружена пустая строка - {current_row} (конец xlsx файла).')
                # empty_row_is_finded = True
        except:
            logger.critical('КРИТИЧЕСКАЯ ОШИБКА: при выполнении фильтрации xlsx файла. Требуется перезапуск программы.')
    return list_with_filtered_values